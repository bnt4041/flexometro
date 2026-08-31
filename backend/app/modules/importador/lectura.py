"""Leer una hoja: CSV o Excel, a filas de texto.

Todo llega como texto y se convierte después (`destinos.convertir`). Fiarse
del tipo que trae la hoja es una trampa conocida: Excel guarda un NIF que
empieza por cero como número y se come el cero, y un código postal de Madrid
pasa de «28001» a 28001. Leyendo todo como texto eso no puede pasar.
"""

import csv
import io

from openpyxl import load_workbook

#: Tope de filas. No es un límite técnico sino de sensatez: la hoja entera se
#: guarda como instantánea para que lo que se importa sea exactamente lo que
#: se previsualizó, y una fila JSONB de medio millón de líneas no es eso, es
#: un problema.
MAX_FILAS = 5000


class ArchivoInvalido(Exception):
    pass


def _texto(valor) -> str:
    if valor is None:
        return ""
    # Excel devuelve los enteros como float: 28001 llega como 28001.0 y un
    # código postal con «.0» detrás no le sirve a nadie.
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    return str(valor).strip()


def leer(nombre: str, contenido: bytes) -> tuple[list[str], list[dict[str, str]]]:
    """Devuelve `(columnas, filas)`. La primera línea son las cabeceras."""
    if nombre.lower().endswith((".xlsx", ".xlsm")):
        return _leer_excel(contenido)
    if nombre.lower().endswith((".csv", ".txt")):
        return _leer_csv(contenido)
    raise ArchivoInvalido("Solo se admiten ficheros .csv, .xlsx o .xlsm")


def _leer_excel(contenido: bytes) -> tuple[list[str], list[dict[str, str]]]:
    try:
        libro = load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise ArchivoInvalido(f"No se ha podido abrir el Excel: {exc}") from exc

    hoja = libro.worksheets[0]
    filas_crudas = hoja.iter_rows(values_only=True)
    try:
        cabeceras = [_texto(c) for c in next(filas_crudas)]
    except StopIteration:
        raise ArchivoInvalido("La hoja está vacía") from None

    columnas = _columnas(cabeceras)
    filas = []
    for cruda in filas_crudas:
        valores = [_texto(c) for c in cruda]
        if not any(valores):
            # Fila en blanco: Excel arrastra cientos al final del archivo.
            continue
        filas.append(dict(zip(columnas, valores + [""] * len(columnas), strict=False)))
        if len(filas) >= MAX_FILAS:
            break
    libro.close()
    return columnas, filas


def _leer_csv(contenido: bytes) -> tuple[list[str], list[dict[str, str]]]:
    # Se prueban dos codificaciones: UTF-8 es lo correcto, pero un CSV
    # exportado desde un Excel español viene en cp1252 y con UTF-8 estricto
    # reventaría en la primera eñe.
    texto = None
    for codificacion in ("utf-8-sig", "cp1252"):
        try:
            texto = contenido.decode(codificacion)
            break
        except UnicodeDecodeError:
            continue
    if texto is None:
        raise ArchivoInvalido("No se ha podido leer el fichero: codificación desconocida")

    try:
        # El delimitador se detecta: en España el CSV de Excel usa `;` porque
        # la coma es el separador decimal.
        dialecto = csv.Sniffer().sniff(texto[:4096], delimiters=";,\t|")
    except csv.Error:
        dialecto = csv.excel
        dialecto.delimiter = ";"

    lector = csv.reader(io.StringIO(texto), dialecto)
    try:
        columnas = _columnas([c.strip() for c in next(lector)])
    except StopIteration:
        raise ArchivoInvalido("El fichero está vacío") from None

    filas = []
    for cruda in lector:
        valores = [c.strip() for c in cruda]
        if not any(valores):
            continue
        filas.append(dict(zip(columnas, valores + [""] * len(columnas), strict=False)))
        if len(filas) >= MAX_FILAS:
            break
    return columnas, filas


def _columnas(cabeceras: list[str]) -> list[str]:
    """Nombres únicos y no vacíos.

    Una hoja real trae columnas sin título y columnas repetidas. Sin esto,
    dos columnas «Teléfono» se pisarían al montar el diccionario de la fila y
    una de las dos se perdería en silencio."""
    vistas: dict[str, int] = {}
    salida = []
    for i, cabecera in enumerate(cabeceras):
        nombre = cabecera or f"Columna {i + 1}"
        if nombre in vistas:
            vistas[nombre] += 1
            nombre = f"{nombre} ({vistas[nombre]})"
        else:
            vistas[nombre] = 1
        salida.append(nombre)
    return salida


def sugerir_mapeo(columnas: list[str], campos) -> dict[str, str]:
    """Empareja columnas con campos por parecido del nombre.

    Ahorra el 90% del trabajo en una hoja normal —«Razón social» encuentra
    `razon_social` sola— y lo que no acierte se corrige a mano. Se compara sin
    tildes ni signos porque nadie escribe la cabecera igual dos veces.
    """
    import unicodedata

    def normalizar(texto: str) -> str:
        sin_tildes = "".join(
            c for c in unicodedata.normalize("NFD", texto.lower())
            if unicodedata.category(c) != "Mn"
        )
        return "".join(c for c in sin_tildes if c.isalnum())

    indice = {normalizar(c): c for c in columnas}
    mapeo = {}
    for campo in campos:
        for candidato in (campo.nombre, campo.etiqueta):
            columna = indice.get(normalizar(candidato))
            if columna:
                mapeo[campo.nombre] = columna
                break
    return mapeo
