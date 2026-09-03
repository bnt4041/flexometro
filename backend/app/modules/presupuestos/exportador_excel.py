"""Exportación del presupuesto a Excel.

Un libro con una hoja "Presupuesto" siempre presente (columnas según lo que
se pida) y, opcionalmente, "Mediciones" y "Descompuestos" — el mismo
contenido que tenían los PDFs fijos que se han retirado, pero en `.xlsx`.
"""

from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.ext.asyncio import AsyncSession

from app.modules.presupuestos import datos_exportacion as datos
from app.modules.presupuestos.models_presupuesto import Presupuesto

_NEGRITA = Font(bold=True)
_ENCABEZADO = Font(bold=True, color="FFFFFF")
_RELLENO_CAPITULO = Font(bold=True)


def _numero(valor: Decimal | None) -> float | None:
    return float(valor) if valor is not None else None


def _autoancho(hoja: Worksheet) -> None:
    for columna in hoja.columns:
        letra = columna[0].column_letter
        ancho = max((len(str(c.value)) for c in columna if c.value is not None), default=8)
        hoja.column_dimensions[letra].width = min(max(ancho + 2, 10), 60)


async def generar_excel(
    session: AsyncSession,
    presupuesto: Presupuesto,
    *,
    incluir_coste: bool = True,
    incluir_venta: bool = False,
    incluir_descompuestos: bool = False,
    incluir_mediciones: bool = False,
    incluir_descripcion: bool = False,
) -> bytes:
    libro = Workbook()
    _hoja_presupuesto(
        libro.active,
        *await datos.capitulos_y_partidas_planos(
            session, presupuesto.id, con_mediciones=incluir_mediciones
        ),
        incluir_coste=incluir_coste,
        incluir_venta=incluir_venta,
        incluir_descripcion=incluir_descripcion,
    )
    libro.active.title = "Presupuesto"

    if incluir_mediciones:
        _, partidas = await datos.capitulos_y_partidas_planos(
            session, presupuesto.id, con_mediciones=True
        )
        _hoja_mediciones(libro.create_sheet("Mediciones"), partidas)

    if incluir_descompuestos:
        conceptos = await datos.conceptos_del_presupuesto(session, presupuesto.id)
        _hoja_descompuestos(libro.create_sheet("Descompuestos"), conceptos)

    buffer = BytesIO()
    libro.save(buffer)
    return buffer.getvalue()


def _hoja_presupuesto(
    hoja: Worksheet,
    capitulos: list[datos.CapituloPlano],
    partidas: list[datos.PartidaPlana],
    *,
    incluir_coste: bool,
    incluir_venta: bool,
    incluir_descripcion: bool,
) -> None:
    encabezados = ["Código", "Resumen", "Unidad", "Medición"]
    if incluir_descripcion:
        encabezados.append("Descripción")
    if incluir_coste:
        encabezados += ["Precio coste", "Importe coste"]
    if incluir_venta:
        encabezados += ["Precio venta", "Importe venta"]
    hoja.append(encabezados)
    for celda in hoja[1]:
        celda.font = _ENCABEZADO

    partidas_por_capitulo: dict[str, list[datos.PartidaPlana]] = {}
    for partida in partidas:
        partidas_por_capitulo.setdefault(partida.capitulo_codigo, []).append(partida)

    for capitulo in capitulos:
        fila = [capitulo.codigo, capitulo.resumen, "", ""]
        if incluir_descripcion:
            fila.append("")
        if incluir_coste:
            fila += ["", _numero(capitulo.importe)]
        if incluir_venta:
            fila += ["", _numero(capitulo.importe_venta)]
        hoja.append(fila)
        for celda in hoja[hoja.max_row]:
            celda.font = _NEGRITA

        for partida in partidas_por_capitulo.get(capitulo.codigo, []):
            fila = [partida.codigo, partida.resumen, partida.unidad, _numero(partida.medicion)]
            if incluir_descripcion:
                fila.append(partida.texto)
            if incluir_coste:
                fila += [_numero(partida.precio), _numero(partida.importe)]
            if incluir_venta:
                fila += [_numero(partida.precio_venta), _numero(partida.importe_venta)]
            hoja.append(fila)

    _autoancho(hoja)


def _hoja_mediciones(hoja: Worksheet, partidas: list[datos.PartidaPlana]) -> None:
    hoja.append(
        ["Capítulo", "Partida", "Comentario", "Uds", "Longitud", "Anchura", "Altura", "Parcial"]
    )
    for celda in hoja[1]:
        celda.font = _ENCABEZADO

    for partida in partidas:
        if not partida.lineas_medicion:
            continue
        hoja.append([f"{partida.capitulo_codigo} · {partida.resumen}", partida.codigo])
        for celda in hoja[hoja.max_row]:
            celda.font = _NEGRITA
        for linea in partida.lineas_medicion:
            hoja.append(
                [
                    "",
                    "",
                    linea.comentario or "",
                    _numero(linea.uds),
                    _numero(linea.longitud),
                    _numero(linea.anchura),
                    _numero(linea.altura),
                    _numero(linea.parcial),
                ]
            )
        hoja.append(["", "", "", "", "", "", "Total", _numero(partida.medicion)])
        for celda in hoja[hoja.max_row]:
            celda.font = _NEGRITA
            celda.alignment = Alignment(horizontal="right")

    _autoancho(hoja)


def _hoja_descompuestos(hoja: Worksheet, conceptos: list) -> None:
    hoja.append(["Código", "Resumen", "Unidad", "Rendimiento", "Factor", "Precio", "Importe"])
    for celda in hoja[1]:
        celda.font = _ENCABEZADO

    for concepto in conceptos:
        hoja.append([concepto.codigo, concepto.resumen, concepto.unidad, "", "", "", _numero(concepto.precio)])
        for celda in hoja[hoja.max_row]:
            celda.font = _NEGRITA
        # `lineas_informe` son `LineaOut` (service.lineas_de): ya vienen con
        # los datos del hijo aplanados (hijo_resumen/hijo_unidad/...), no un
        # objeto `.hijo` anidado, y el precio del componente es
        # `hijo_precio` (LineaOut no tiene `.precio` propio).
        for linea in getattr(concepto, "lineas_informe", []):
            hoja.append(
                [
                    "",
                    linea.hijo_resumen,
                    linea.hijo_unidad,
                    _numero(linea.rendimiento),
                    _numero(linea.factor),
                    _numero(linea.hijo_precio),
                    _numero(linea.rendimiento * linea.factor * linea.hijo_precio),
                ]
            )

    _autoancho(hoja)
